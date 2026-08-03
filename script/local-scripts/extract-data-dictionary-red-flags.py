"""Build a clean report of Roxanne's red-flagged field changes."""
from openpyxl import load_workbook

XLSX = "/Users/raby.whyte/code/hmpps-accredited-programmes-api/doc/2026.07.08_copy_Probation Digital Data review December 251.xlsx"
wb = load_workbook(XLSX, data_only=True)
ws = wb["Accredited Programmes Custody"]

def red_ish(rgb):
    if not rgb or not isinstance(rgb, str) or len(rgb) < 6: return False
    r, g, b = int(rgb[-6:-4], 16), int(rgb[-4:-2], 16), int(rgb[-2:], 16)
    return r >= 150 and g < 100 and b < 100

def is_red(cell):
    fc = cell.font.color
    return fc and fc.type == 'rgb' and red_ish(fc.rgb)

# find rows that have ANY red cell in columns F-I (the decision columns)
DECISION_COLS = [6, 7, 8, 9]  # F, G, H, I
red_rows = set()
for r in range(14, ws.max_row + 1):
    for cc in DECISION_COLS:
        if is_red(ws.cell(row=r, column=cc)):
            red_rows.add(r)
            break

# for each red row, propagate to find the section header (last non-empty A above)
def entity_at(r):
    for rr in range(r, 13, -1):
        v = ws.cell(row=rr, column=1).value
        if v:
            return str(v).strip()
    return "?"

print(f"Total red-flagged data rows: {len(red_rows)}\n")

# structured output grouped by entity
from collections import defaultdict
by_entity = defaultdict(list)
for r in sorted(red_rows):
    ent = entity_at(r)
    elem = ws.cell(row=r, column=3).value or ""
    desc = ws.cell(row=r, column=4).value or ""
    ex = ws.cell(row=r, column=5).value or ""
    sar = ws.cell(row=r, column=7).value or ""    # SAR data Y/N
    in_api = ws.cell(row=r, column=8).value or "" # In SAR API Y/N/NA
    notes = ws.cell(row=r, column=9).value or ""
    by_entity[ent].append({
        "row": r, "element": str(elem).strip(), "description": str(desc).strip()[:120],
        "example": str(ex).strip()[:60],
        "sar_data": str(sar).strip(), "in_api": str(in_api).strip(),
        "notes": str(notes).strip()[:180],
        "h_red": is_red(ws.cell(row=r, column=8)),
        "i_red": is_red(ws.cell(row=r, column=9)),
    })

for ent, rows in by_entity.items():
    print(f"\n### ENTITY: {ent}  ({len(rows)} flagged rows)")
    for row in rows:
        marker = ""
        if row["h_red"]: marker += " [H-red]"
        if row["i_red"]: marker += " [I-red]"
        print(f"  R{row['row']:>3} {row['element']:<45} sar={row['sar_data']:<6} in_api={row['in_api']:<6}{marker}")
        print(f"        desc: {row['description']}")
        if row["notes"]:
            print(f"        note: {row['notes']}")

