#!/usr/bin/env python3
"""
Sweep the Additional Notes column on the "Accredited Programmes Custody"
sheet of the Data Dictionary spreadsheet and print every row with a
non-empty note.

Why: our earlier DD cross-checks (PRs 1-3 cross-check commits) filtered
on Roxanne's red-flagged rows only. Dev-authored notes on non-red rows
(e.g. row 88 `programme_pathway` "should be on the report", row 165
`original_referral_id` "do not add the uuid") were missed and later
surfaced only because Deborah's DM prompted a wider look. This script
is the belt-and-braces: read every note-bearing row, not just the red
ones, so future DD refreshes cross-check the whole surface.

Usage:
  python3 doc/planning/APG-2546/scripts/dd-notes-sweep.py [PATH_TO_XLSX]

Defaults to the local working copy path (untracked in git):
  doc/Copy of 2026.07.08_copy_Probation Digital Data review December 251.xlsx

Filter to interesting rows:
  python3 doc/planning/APG-2546/scripts/dd-notes-sweep.py | grep -B1 '10.07'
  python3 doc/planning/APG-2546/scripts/dd-notes-sweep.py | grep -B1 -v 'Raby 29.07'

Dependencies:
  pip install openpyxl

Sheet layout (assumed, verified 2026-08-04 pm):
  Row 13 = header
  Col A = Entity (table name; on non-first row of a table, blank)
  Col C = Element (field name)
  Col F = Mandatory Y/N
  Col G = SAR data Y/N
  Col H = In SAR API Y/N
  Col I = Additional Notes (the free-text column we're mining)

Kept alongside the planning docs so a fresh agent picking up an APG-2546-
adjacent ticket can re-run without having to rebuild the script. Also
serves as a reference for how to read the DD without accidentally
filtering on colour formatting alone.
"""

import sys
from pathlib import Path

import openpyxl

DEFAULT_XLSX = (
    "doc/Copy of 2026.07.08_copy_"
    "Probation Digital Data review December 251.xlsx"
)
SHEET = "Accredited Programmes Custody"

# Columns (1-based).
COL_ENTITY = 1
COL_ELEMENT = 3
COL_MANDATORY = 6
COL_SAR_DATA = 7
COL_IN_API = 8
COL_NOTES = 9

# Data rows start after the header at row 13; row 14 is the "Example"
# instructional row. Start sweeping from row 15.
FIRST_DATA_ROW = 15


def sweep(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {SHEET!r} not found in {xlsx_path}. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET]

    current_entity = None
    hits = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        entity_cell = ws.cell(r, COL_ENTITY).value
        element_cell = ws.cell(r, COL_ELEMENT).value

        # Entity name only appears on the first row of a table
        # (when Element is blank). Any Element-less row with an Entity
        # is a header row; carry it forward.
        if entity_cell and not element_cell:
            current_entity = entity_cell
            continue
        if entity_cell:
            current_entity = entity_cell
        if not element_cell:
            continue

        note = ws.cell(r, COL_NOTES).value
        if not note:
            continue

        hits.append(
            (
                r,
                current_entity,
                element_cell,
                ws.cell(r, COL_MANDATORY).value,
                ws.cell(r, COL_SAR_DATA).value,
                ws.cell(r, COL_IN_API).value,
                note,
            )
        )

    print(f"TOTAL rows with notes: {len(hits)}")
    print()
    for r, entity, element, mand, sar_data, in_api, note in hits:
        print(
            f"[row {r}] {entity} . {element}  "
            f"| Mand={mand} SAR={sar_data} API={in_api}"
        )
        for line in str(note).splitlines():
            print(f"    NOTE: {line}")
        print()


def main() -> int:
    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX)
    if not xlsx.exists():
        raise SystemExit(
            f"Spreadsheet not found: {xlsx}\n"
            "The working copy of the DD is intentionally untracked in "
            "git — either pass an explicit path or drop your copy at "
            f"{DEFAULT_XLSX}."
        )
    sweep(xlsx)
    return 0


if __name__ == "__main__":
    sys.exit(main())

