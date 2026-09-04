#!/usr/bin/env python3
"""
Apply the APG-2546 rounds 1 + 2 "In SAR API - Y/N or N/A" (column H)
updates to the Data Dictionary spreadsheet, sheet
"Accredited Programmes Custody".

Baseline input:
  ~/Downloads/Copy of 2026.07.08_copy_Probation Digital Data
  review December 251.xlsx
  (the copy Roxanne distributed 2026-07-08; untracked in git per
  repo convention.)

Output:
  Same directory, filename suffixed
  "..._APG-2546-round-2-update.xlsx".
  Safe to attach to the reply to Roxanne (2026-08-19 email).

Every column-H change is derived from a specific merged PR on
`origin/main`; each entry in DELTAS carries the PR number as its
reason string, so Roxanne (or any reviewer) can trace a row change
back to a merge SHA.

Rounds 1 + 2 code merges — all on `origin/main @ 99264496`:
  PR-1  #1107 50f67cff  remove auditRecords
  PR-2  #1109 cd306c99  remove referralStatusHistory + referralStatusReasons
  PR-3  #1110 d6587351  remove sexualOffenceDetails + selectedSexualOffenceDetails
  PR-4  #1111 2a79b856  strip pniResultId + oasysAssessmentId from oasysPniResults
  PR-5  #1112 50968d07  strip SarPerson.id + SarOrganisation.id + SarReferral.originalReferralId
  PR-7  #1113 baee4510  strip SarOriginalReferral.id UUID
  PR-8  #1116 b7b05283  remove PNI + OASys-PNI + Person sections
  PR-9  #1117 f8e04ab0  scrub prisonerNumber from surviving SAR sections
  PR-10 #1118 d710fa7f  fold organisation into referral; drop top-level organisations[]
  PR-11 #1119 47488c8a  drop top-level staff[]; retain inline POM surnames
  PR-12 #1120 99264496  hygiene / fixture widening (no DTO shape change)

DELTAS covers every row where the column-H value in the July 8
baseline diverges from the code truth on `origin/main @ 99264496`.
Rows already correct in the baseline (e.g. audit_record.audit_record_id
already No, pni_result.pni_result_id already No via Roxanne's 10.07
annotations) are NOT included — the point is the column-H delta,
not a re-flip of already-correct values.

Also includes the two "DD drift" corrections flagged in the
DELIVERY-LOG (rows 109 + 224) — code truth vs baseline.

Run:
  python3 doc/planning/APG-2546/scripts/dd-column-h-update.py
"""
import os
import shutil
import sys
from openpyxl import load_workbook

HOME = os.path.expanduser("~")
BASELINE = os.path.join(
    HOME,
    "Downloads",
    "Copy of 2026.07.08_copy_Probation Digital Data review December 251.xlsx",
)
OUTPUT = os.path.join(
    HOME,
    "Downloads",
    "Copy of 2026.07.08_copy_Probation Digital Data review December 251_APG-2546-round-2-update.xlsx",
)

SHEET = "Accredited Programmes Custody"
COL_H = 8  # "In SAR API - Y/N or N/A"

# Every row where baseline column-H diverges from origin/main @ 99264496.
# (row, expected_baseline_H, new_H, entity.element, reason)
DELTAS = [
    # --- PR-1 remove auditRecords (whole Content.auditRecords section gone) ---
    (22, "Yes", "No", "audit_record.prison_number",       "PR-1 (#1107) removed Content.auditRecords"),
    (23, "Yes", "No", "audit_record.referrer_username",   "PR-1 (#1107) removed Content.auditRecords"),
    (24, "Yes", "No", "audit_record.referral_status_from","PR-1 (#1107) removed Content.auditRecords"),
    (25, "Yes", "No", "audit_record.referral_status_to",  "PR-1 (#1107) removed Content.auditRecords"),
    (27, "Yes", "No", "audit_record.course_name",         "PR-1 (#1107) removed Content.auditRecords"),
    (28, "Yes", "No", "audit_record.course_location",     "PR-1 (#1107) removed Content.auditRecords"),
    (29, "Yes", "No", "audit_record.audit_action",        "PR-1 (#1107) removed Content.auditRecords"),
    (30, "Yes", "No", "audit_record.audit_username",      "PR-1 (#1107) removed Content.auditRecords"),
    (31, "Yes", "No", "audit_record.audit_date_time",     "PR-1 (#1107) removed Content.auditRecords"),

    # --- PR-9 scrub prisonerNumber from surviving SAR sections ---
    (47, "Yes", "No", "course_participation.prison_number","PR-9 (#1117) scrubbed prisonerNumber"),

    # --- PR-8 whole oasys_pni_result section removed ---
    # (PR-4 first stripped pni_result_id + oasys_assessment_id in round 1;
    #  PR-8 then removed the whole section in round 2. Net = all four rows No.)
    (85, "Yes", "No", "oasys_pni_result.pni_result_id",           "PR-8 (#1116) removed oasysPniResults section (PR-4 also stripped)"),
    (86, "Yes", "No", "oasys_pni_result.prison_number",           "PR-8 (#1116) removed oasysPniResults section"),
    (87, "Yes", "No", "oasys_pni_result.oasys_assessment_id",     "PR-8 (#1116) removed oasysPniResults section (PR-4 also stripped)"),
    (88, "Yes", "No", "oasys_pni_result.programme_pathway",       "PR-8 (#1116) removed oasysPniResults section"),

    # --- PR-5 stripped organisation.organisation_id from SarOrganisation ---
    # --- PR-10 removed the whole top-level organisations[] block, folded
    #     `name` inline on referral as `organisationName` ---
    (105, "Yes", "No",  "organisation.organisation_id",  "PR-5 (#1112) stripped SarOrganisation.id; PR-10 (#1118) then removed top-level organisations[]"),
    (106, "Yes", "No",  "organisation.code",             "PR-10 (#1118) removed top-level organisations[]; code is not surfaced elsewhere"),
    # R107 organisation.name STAYS Yes — now surfaces as
    # `organisationName` inline on each referral (PR-10). No flip.
    (108, "Yes", "No",  "organisation.gender",           "PR-10 (#1118) removed top-level organisations[]; gender is not surfaced elsewhere"),
    (109, "Yes", "No",  "organisation.is_national",      "DD drift correction: code has always been No; APG-2494 won't-do; Q2 closed 2026-08-04 on 'leave off' default"),

    # --- PR-5 stripped SarPerson.id ---
    # --- PR-8 removed the whole Person section in round 2 ---
    (111, "Yes", "No", "person.person_id",                 "PR-5 (#1112) stripped SarPerson.id; PR-8 (#1116) then removed Person section"),
    (112, "Yes", "No", "person.prison_number",             "PR-8 (#1116) removed Person section"),
    (113, "Yes", "No", "person.forename",                  "PR-8 (#1116) removed Person section"),
    (114, "Yes", "No", "person.surname",                   "PR-8 (#1116) removed Person section"),
    (115, "Yes", "No", "person.conditional_release_date",  "PR-8 (#1116) removed Person section"),
    (116, "Yes", "No", "person.parole_eligibility_date",   "PR-8 (#1116) removed Person section"),
    (117, "Yes", "No", "person.tariff_expiry_date",        "PR-8 (#1116) removed Person section"),
    (118, "Yes", "No", "person.earliest_release_date",     "PR-8 (#1116) removed Person section"),
    (119, "Yes", "No", "person.earliest_release_date_type","PR-8 (#1116) removed Person section"),
    (120, "Yes", "No", "person.indeterminate_sentence",    "PR-8 (#1116) removed Person section"),
    (121, "Yes", "No", "person.non_dto_release_date_type", "PR-8 (#1116) removed Person section"),
    (122, "Yes", "No", "person.sentence_type",             "PR-8 (#1116) removed Person section"),
    (123, "Yes", "No", "person.location",                  "PR-8 (#1116) removed Person section"),
    (124, "Yes", "No", "person.gender",                    "PR-8 (#1116) removed Person section"),

    # --- PR-8 removed whole PNI results section in round 2 ---
    # (R127, 128, 131 already flipped to No by Roxanne 10.07 - no delta needed.)
    (129, "Yes", "No", "pni_result.prison_number",                  "PR-8 (#1116) removed pniResults section"),
    (130, "Yes", "No", "pni_result.crn",                            "PR-8 (#1116) removed pniResults section"),
    (132, "Yes", "No", "pni_result.oasys_assessment_completed_date","PR-8 (#1116) removed pniResults section"),
    (133, "Yes", "No", "pni_result.programme_pathway",              "PR-8 (#1116) removed pniResults section"),
    (134, "Yes", "No", "pni_result.needs_classification",           "PR-8 (#1116) removed pniResults section"),
    (135, "Yes", "No", "pni_result.overall_needs_score",            "PR-8 (#1116) removed pniResults section"),
    (136, "Yes", "No", "pni_result.risk_classification",            "PR-8 (#1116) removed pniResults section"),
    (137, "Yes", "No", "pni_result.pni_assessment_date",            "PR-8 (#1116) removed pniResults section"),
    (138, "Yes", "No", "pni_result.pni_valid",                      "PR-8 (#1116) removed pniResults section"),
    # R139 pni_result_json — Roxanne flipped to Yes 10.07 with note
    # "these are in SAR report hence H should be Yes. Updated".
    # SUPERSEDED by Deborah's 2026-08-13 review-meeting decision:
    # PNI data now sourced by SAR consumers via ARNs Probation Hub,
    # so replicating in ACP SAR is duplicative. Recorded in
    # ROUND-2-PLAN §"DD spreadsheet override".
    (139, "Yes", "No", "pni_result.pni_result_json",                "PR-8 (#1116) removed pniResults section — supersedes Roxanne's 10.07 flip per Deborah's 2026-08-13 decision (PNI now via ARNs Probation Hub)"),
    (140, "Yes", "No", "pni_result.basic_skills_score",             "PR-8 (#1116) removed pniResults section"),

    # --- PR-9 scrub prisonerNumber from surviving SAR sections ---
    (153, "Yes", "No", "referral.prison_number", "PR-9 (#1117) scrubbed prisonerNumber from Referrals"),

    # --- PR-5 + PR-7 stripped original_referral_id UUID ---
    (165, "Yes", "No", "referral.original_referral_id",
     "PR-5 (#1112) stripped SarReferral.originalReferralId; PR-7 (#1113) also stripped the nested SarOriginalReferral.id — resolved originalReferral sub-block (courseName, submittedOn, referrerSurname, organisationName, etc.) retained without any UUID"),

    # --- PR-2 remove referralStatusHistory + referralStatusReasons ---
    (192, "Yes", "No", "referral_status_history.status_history_id",  "PR-2 (#1109) removed Content.referralStatusHistory"),
    (193, "Yes", "No", "referral_status_history.referral_id",        "PR-2 (#1109) removed Content.referralStatusHistory"),
    (194, "Yes", "No", "referral_status_history.status",             "PR-2 (#1109) removed Content.referralStatusHistory"),
    (195, "Yes", "No", "referral_status_history.previous_status",    "PR-2 (#1109) removed Content.referralStatusHistory"),
    (196, "Yes", "No", "referral_status_history.category",           "PR-2 (#1109) removed Content.referralStatusHistory"),
    (197, "Yes", "No", "referral_status_history.reason",             "PR-2 (#1109) removed Content.referralStatusHistory"),
    (198, "Yes", "No", "referral_status_history.notes",              "PR-2 (#1109) removed Content.referralStatusHistory"),
    (199, "Yes", "No", "referral_status_history.status_start_date",  "PR-2 (#1109) removed Content.referralStatusHistory"),
    (200, "Yes", "No", "referral_status_history.status_end_date",    "PR-2 (#1109) removed Content.referralStatusHistory"),
    (201, "Yes", "No", "referral_status_history.duration_at_this_status", "PR-2 (#1109) removed Content.referralStatusHistory"),
    (202, "Yes", "No", "referral_status_history.username",           "PR-2 (#1109) removed Content.referralStatusHistory"),

    (205, "Yes", "No", "referral_status_reason.code",                        "PR-2 (#1109) removed Content.referralStatusReasons"),
    (206, "Yes", "No", "referral_status_reason.referral_status_category_code","PR-2 (#1109) removed Content.referralStatusReasons"),
    (207, "Yes", "No", "referral_status_reason.description",                 "PR-2 (#1109) removed Content.referralStatusReasons"),
    (208, "Yes", "No", "referral_status_reason.active",                      "PR-2 (#1109) removed Content.referralStatusReasons"),
    (209, "Yes", "No", "referral_status_reason.deselect_open",               "PR-2 (#1109) removed Content.referralStatusReasons"),

    # --- DD drift correction: referrer_user.referrer_username ---
    (224, "No", "Yes", "referrer_user.referrer_username",
     "DD drift correction: note already says 'Yes if we can provide surname'; we do surface a surname via APG-2492 (referrer.username -> resolved surname)"),

    # --- PR-3 remove sexualOffenceDetails + selectedSexualOffenceDetails ---
    (226, "Yes", "No", "selected_sexual_offence_details.id",                     "PR-3 (#1110) removed Content.selectedSexualOffenceDetails"),
    (227, "Yes", "No", "selected_sexual_offence_details.referral_id",            "PR-3 (#1110) removed Content.selectedSexualOffenceDetails"),
    (228, "Yes", "No", "selected_sexual_offence_details.sexual_offence_details_id","PR-3 (#1110) removed Content.selectedSexualOffenceDetails"),
    (233, "Yes", "No", "sexual_offence_details.id",          "PR-3 (#1110) removed Content.sexualOffenceDetails"),
    (234, "Yes", "No", "sexual_offence_details.category",    "PR-3 (#1110) removed Content.sexualOffenceDetails"),
    (235, "Yes", "No", "sexual_offence_details.description", "PR-3 (#1110) removed Content.sexualOffenceDetails"),
    (237, "Yes", "No", "sexual_offence_details.score",       "PR-3 (#1110) removed Content.sexualOffenceDetails"),
]


def main():
    if not os.path.exists(BASELINE):
        print(f"ERROR: baseline not found at {BASELINE}", file=sys.stderr)
        return 1

    shutil.copy2(BASELINE, OUTPUT)
    wb = load_workbook(OUTPUT)
    ws = wb[SHEET]

    print(f"Baseline: {BASELINE}")
    print(f"Output:   {OUTPUT}")
    print(f"Sheet:    {SHEET}\n")
    print(f"{'Row':<5} {'Field':<50} {'Was':<5} {'Now':<5} Reason")
    print("-" * 120)

    baseline_mismatches = []
    applied = 0
    for row, expected, new, label, reason in DELTAS:
        actual = ws.cell(row, COL_H).value
        actual_norm = str(actual).strip() if actual is not None else ""
        if actual_norm.lower() != expected.lower():
            baseline_mismatches.append((row, label, expected, actual))
            marker = "!"
        else:
            marker = " "
        ws.cell(row, COL_H).value = new
        print(f"{marker}R{row:<3} {label:<50} {expected:<5} -> {new:<5} {reason}")
        applied += 1

    wb.save(OUTPUT)

    print("-" * 120)
    print(f"\nApplied {applied} column-H changes.")
    if baseline_mismatches:
        print(f"\n⚠️  {len(baseline_mismatches)} row(s) diverged from expected baseline value:")
        for row, label, expected, actual in baseline_mismatches:
            print(f"     R{row} {label}: expected {expected!r}, found {actual!r}")
        print("     -> baseline drifted since 2026-07-08 copy? re-verify against Roxanne's copy before sending.")
    else:
        print("Every baseline value matched expectations. Output is safe to send.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

